from __future__ import annotations

import json
import re
import zipfile
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import load_workbook
from pypdf import PdfReader

ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "data"
OUT = DATA / "generated"
OUT.mkdir(parents=True, exist_ok=True)

EXCELS = [DATA / "costeokit ejemplo.xlsx", DATA / "Costeokit Herramienta en blanco.xlsx"]
PDFS = list(DATA.glob("*.pdf"))

NS = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}


def cell_to_row_col(cell_ref: str):
    m = re.match(r"([A-Z]+)([0-9]+)", cell_ref)
    if not m:
        return None, None
    col_s, row_s = m.groups()
    col = 0
    for ch in col_s:
        col = col * 26 + (ord(ch) - ord("A") + 1)
    return int(row_s), col


def column_letter(col: int) -> str:
    out = ""
    while col:
        col, rem = divmod(col - 1, 26)
        out = chr(65 + rem) + out
    return out


def sheet_xml_paths(xlsx: Path):
    with zipfile.ZipFile(xlsx) as z:
        wb = ET.fromstring(z.read("xl/workbook.xml"))
        rels = ET.fromstring(z.read("xl/_rels/workbook.xml.rels"))
        rel_map = {r.attrib["Id"]: r.attrib["Target"] for r in rels}
        result = []
        for sheet in wb.findall("main:sheets/main:sheet", NS):
            name = sheet.attrib["name"]
            rid = sheet.attrib["{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"]
            target = rel_map[rid]
            if not target.startswith("xl/"):
                target = "xl/" + target
            result.append((name, target))
        return result


def xml_sheet_metrics(xlsx: Path):
    metrics = {}
    with zipfile.ZipFile(xlsx) as z:
        for name, path in sheet_xml_paths(xlsx):
            xml = ET.fromstring(z.read(path))
            cells = xml.findall(".//main:c", NS)
            populated = 0
            formulas = []
            values = 0
            stringish = 0
            numeric = 0
            min_row = min_col = 10**9
            max_row = max_col = 0
            for c in cells:
                ref = c.attrib.get("r", "")
                row, col = cell_to_row_col(ref)
                f = c.find("main:f", NS)
                v = c.find("main:v", NS)
                is_pop = f is not None or v is not None or c.attrib.get("t") in {"inlineStr", "s", "str"}
                if is_pop:
                    populated += 1
                    if row and col:
                        min_row, min_col = min(min_row, row), min(min_col, col)
                        max_row, max_col = max(max_row, row), max(max_col, col)
                if f is not None:
                    formulas.append({"cell": ref, "formula": f.text or "", "type": f.attrib.get("t")})
                elif v is not None:
                    values += 1
                    if c.attrib.get("t") in {"s", "str", "inlineStr"}:
                        stringish += 1
                    else:
                        numeric += 1
            metrics[name] = {
                "populated_cells": populated,
                "used_range": None if populated == 0 else f"{column_letter(min_col)}{min_row}:{column_letter(max_col)}{max_row}",
                "formula_count": len(formulas),
                "value_count": values,
                "stringish_count": stringish,
                "numeric_count": numeric,
                "sample_formulas": formulas[:25],
            }
    return metrics


def get_visible_rows(ws, max_rows=80, max_cols=14):
    rows = []
    for r in range(1, min(ws.max_row or 1, max_rows) + 1):
        vals = []
        any_val = False
        for c in range(1, min(ws.max_column or 1, max_cols) + 1):
            v = ws.cell(r, c).value
            if v is not None:
                any_val = True
            vals.append(v)
        if any_val:
            rows.append((r, vals))
    return rows


def normalize_header(v):
    if v is None:
        return ""
    return str(v).strip().replace("\n", " ")


def detect_header_rows(ws, max_rows=60):
    candidates = []
    for r in range(1, min(ws.max_row or 1, max_rows) + 1):
        values = [normalize_header(ws.cell(r, c).value) for c in range(1, min(ws.max_column or 1, 20) + 1)]
        nonempty = [v for v in values if v]
        if len(nonempty) >= 3:
            score = sum(1 for v in nonempty if len(v) <= 45) + len(nonempty)
            candidates.append({"row": r, "nonempty": len(nonempty), "values": nonempty[:12], "score": score})
    return sorted(candidates, key=lambda x: (-x["score"], x["row"]))[:5]


def audit_excel(xlsx: Path):
    formulas_wb = load_workbook(xlsx, read_only=False, data_only=False)
    values_wb = load_workbook(xlsx, read_only=False, data_only=True)
    xml_metrics = xml_sheet_metrics(xlsx)
    sheets = []
    for s in formulas_wb.sheetnames:
        ws = formulas_wb[s]
        vws = values_wb[s]
        metrics = xml_metrics.get(s, {})
        formulas = []
        text_counter = Counter()
        for row in ws.iter_rows():
            for cell in row:
                val = cell.value
                if isinstance(val, str) and val.startswith("="):
                    formulas.append({"cell": cell.coordinate, "formula": val[:500], "cached": vws[cell.coordinate].value})
                elif isinstance(val, str):
                    stripped = val.strip()
                    if stripped:
                        text_counter[stripped[:100]] += 1
        sheets.append({
            "name": s,
            "used_range": metrics.get("used_range"),
            "populated_cells": metrics.get("populated_cells", 0),
            "formula_count_xml": metrics.get("formula_count", 0),
            "formula_count_openpyxl": len(formulas),
            "header_candidates": detect_header_rows(ws),
            "top_texts": [t for t, _ in text_counter.most_common(20)],
            "sample_formulas": formulas[:35] or metrics.get("sample_formulas", []),
        })
    return {"file": str(xlsx.relative_to(ROOT)), "sheets": sheets}


def formula_features(audits):
    refs = Counter()
    funcs = Counter()
    suspicious = []
    formula_by_sheet = defaultdict(int)
    for wb in audits:
        for sh in wb["sheets"]:
            formula_by_sheet[(wb["file"], sh["name"])] = sh["formula_count_openpyxl"] or sh["formula_count_xml"]
            for sf in sh.get("sample_formulas", []):
                f = sf.get("formula", "")
                for func in re.findall(r"\b([A-Z][A-Z0-9_.]*)\s*\(", f):
                    funcs[func] += 1
                for ref in re.findall(r"'([^']+)'!|([A-Za-zÁÉÍÓÚÜÑáéíóúüñ ]+)!", f):
                    name = ref[0] or ref[1]
                    if name:
                        refs[name] += 1
                if any(term in f.upper() for term in ["BASE DE DATOS", "$I$", "MERMA", "CONSUMO", "PRECIO", "COST"]):
                    suspicious.append({"workbook": wb["file"], "sheet": sh["name"], **sf})
    return {
        "top_functions_in_samples": funcs.most_common(30),
        "top_sheet_refs_in_samples": refs.most_common(30),
        "formula_counts_by_sheet": [{"workbook": k[0], "sheet": k[1], "count": v} for k, v in sorted(formula_by_sheet.items())],
        "business_relevant_sample_formulas": suspicious[:120],
    }


def extract_pdf(pdf: Path):
    reader = PdfReader(str(pdf))
    pages = []
    full = []
    for i, page in enumerate(reader.pages, start=1):
        text = page.extract_text() or ""
        full.append(text)
        lines = [re.sub(r"\s+", " ", line).strip() for line in text.splitlines()]
        lines = [l for l in lines if l]
        pages.append({"page": i, "preview": " ".join(lines)[:1200]})
    text = "\n".join(full)
    rules = []
    keywords = [
        "food cost", "merma", "factor de correccion", "margen", "punto de equilibrio",
        "costo teorico", "costo real", "estandarizacion", "sub-recetas", "inventario",
        "pedidos", "ingresos", "gastos", "analisis de precios", "errores comunes"
    ]
    lower = text.lower()
    for kw in keywords:
        idx = lower.find(kw)
        if idx >= 0:
            snippet = re.sub(r"\s+", " ", text[max(0, idx-250): idx+900]).strip()
            rules.append({"keyword": kw, "snippet": snippet})
    return {"file": str(pdf.relative_to(ROOT)), "pages": len(reader.pages), "pages_preview": pages, "keyword_snippets": rules}


def build_markdown(audits, features, pdfs):
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    lines = []
    lines += ["# CosteoKit — Auditoría técnica de fuentes", "", f"**Generado:** {now}", ""]
    lines += ["## 1. Fuentes auditadas", ""]
    for wb in audits:
        lines.append(f"- `{wb['file']}` — {len(wb['sheets'])} hojas")
    for p in pdfs:
        lines.append(f"- `{p['file']}` — {p['pages']} páginas")
    lines += ["", "## 2. Lectura arquitectónica", ""]
    lines += [
        "La app debe tratar estos archivos como **fuentes de conocimiento de negocio**, no como una especificación técnica perfecta.",
        "El Excel trae estructura, nombres, fórmulas y ejemplos; el PDF trae el porqué operativo: food cost, merma, margen, punto de equilibrio, inventario y errores comunes.",
        "La implementación correcta es separar el dominio en entidades y servicios testeables, y usar el importador sólo como puente inicial de datos.",
        "",
    ]
    lines += ["## 3. Inventario de hojas Excel", ""]
    for wb in audits:
        lines.append(f"### `{wb['file']}`")
        lines.append("")
        lines.append("| Hoja | Rango usado | Celdas pobladas | Fórmulas | Lectura funcional probable |")
        lines.append("|---|---:|---:|---:|---|")
        for sh in wb["sheets"]:
            role = infer_role(sh["name"])
            lines.append(f"| {sh['name']} | {sh.get('used_range') or '-'} | {sh.get('populated_cells') or 0} | {sh.get('formula_count_openpyxl') or sh.get('formula_count_xml') or 0} | {role} |")
        lines.append("")
    lines += ["## 4. Mapeo hoja → módulo de app", ""]
    mapping = [
        ("Configuracion", "business_settings", "Configuración del negocio: moneda, impuestos, margen objetivo, food cost máximo."),
        ("Base de Datos", "suppliers, ingredients", "Catálogo de insumos, proveedores, unidades, precios, consumos y mermas/factores."),
        ("Inventario", "inventory_counts, inventory_count_lines", "Conteos físicos y stock actual por insumo."),
        ("Pedidos", "purchase_suggestions, purchase_suggestion_lines", "Pedido sugerido por consumo máximo, stock y proveedor."),
        ("Ingresos y Gastos", "monthly_ledgers, monthly_ledger_lines", "Registro mensual/semanal de ventas y costos."),
        ("Punto de Equilibrio", "break_even service/report", "Cálculo de ventas/unidades necesarias para no perder plata."),
        ("Analisis de Precios", "pricing read model/report", "Comparación de costo, precio, food cost y margen."),
        ("Indice de Recetas", "recipes summary/read model", "Listado de recetas con costos/estado; NO fuente de verdad de costos."),
        ("Sub-recetas", "sub_recipes, sub_recipe_lines", "Preparaciones reutilizables con costo por unidad de salida."),
        ("Receta 01..20", "recipes, recipe_lines", "Recetas individuales: líneas de insumos/sub-recetas, rendimiento, precio sugerido."),
        ("Instrucciones", "onboarding/help", "Material de ayuda; útil para microcopy y validaciones, no para cálculo."),
    ]
    lines.append("| Hoja | Entidades/módulo | Implementación |")
    lines.append("|---|---|---|")
    for a,b,c in mapping:
        lines.append(f"| {a} | `{b}` | {c} |")
    lines += ["", "## 5. Reglas de negocio extraídas del PDF", ""]
    if pdfs:
        for rule in summarize_pdf_rules(pdfs[0]):
            lines.append(f"- **{rule[0]}:** {rule[1]}")
    lines += ["", "## 6. Hallazgos técnicos del Excel", ""]
    lines += [
        "- Ambos Excel tienen las mismas 30 hojas esperadas: configuración, base de datos, inventario, pedidos, finanzas, equilibrio, pricing, índice, sub-recetas y 20 recetas.",
        "- Hay que diferenciar el Excel de ejemplo del Excel en blanco: el primero sirve para entender datos y resultados esperados; el segundo para detectar estructura y defaults.",
        "- Los costos visibles en hojas agregadas como `Indice de Recetas` o `Analisis de Precios` deben tratarse como reportes derivados, no como fuente primaria.",
        "- Las recetas deben recalcularse desde `ingredients`, `sub_recipes` y `recipe_lines`; no desde valores cacheados del workbook.",
        "- Las unidades y conversiones deben ser explícitas. Si una receta usa gramos y el insumo se compra por kg, el motor debe convertir; no esconderlo en fórmulas sueltas.",
        "- Consumo diario, merma y factor de corrección deben ser campos distintos. Si el Excel mezcla conceptos, la app debe corregir el modelo.",
        "",
    ]
    lines += ["## 7. Motor de cálculo mínimo", ""]
    engine = [
        "`convertUnit(value, fromUnit, toUnit)`",
        "`calculateIngredientUnitCost(ingredient)`",
        "`calculateRecipeLineCost(line, context)`",
        "`calculateSubRecipeCost(subRecipe, context)`",
        "`calculateRecipeCost(recipe, context)`",
        "`calculateSuggestedPrice(cost, targetMarginPct, taxPct)`",
        "`calculateFoodCostPct(cost, priceWithoutTax)`",
        "`calculateGrossMarginPct(cost, priceWithoutTax)`",
        "`calculateInventoryReorderPoint(maxDailyConsumption)`",
        "`calculatePurchaseSuggestion(inventoryLine, ingredient)`",
        "`calculateBreakEven(input)`",
    ]
    for e in engine:
        lines.append(f"- {e}")
    lines += ["", "## 8. Tests obligatorios antes de UI pesada", ""]
    tests = [
        "Conversión kg ↔ gramo, L ↔ ml, docena ↔ unidad.",
        "Costo unitario de insumo con precio de compra, cantidad, unidad de compra, unidad de uso y merma.",
        "Costo de línea de receta usando insumo directo.",
        "Costo de sub-receta y costo por unidad de salida usable.",
        "Costo total y unitario de receta según porciones/rendimiento.",
        "Precio sugerido con margen objetivo e IVA.",
        "Food cost y margen bruto.",
        "Semáforo de pricing: OK / ADJUST / REVIEW.",
        "Punto de pedido: `maxDailyConsumption * 3`.",
        "Punto de equilibrio mensual/diario.",
        "Importador: detecta 30 hojas, estructura esperada y reporta inconsistencias sin confiar en cacheados.",
    ]
    for t in tests:
        lines.append(f"- {t}")
    lines += ["", "## 9. Estrategia de importación", ""]
    lines += [
        "1. Leer workbook y validar presencia de hojas esperadas.",
        "2. Parsear configuración e insumos primero.",
        "3. Parsear sub-recetas antes de recetas, porque pueden ser dependencia de recetas.",
        "4. Parsear recetas 01..20 como entidades normalizadas, no como hojas permanentes.",
        "5. Recalcular costos con `CostingEngine` propio.",
        "6. Comparar resultados recalculados contra valores visibles del Excel sólo como diagnóstico.",
        "7. Generar `excel_imports.report_json` con warnings, errores y decisiones de mapeo.",
        "8. Confirmar importación; desde ese punto la app es fuente de verdad.",
        "",
    ]
    lines += ["## 10. Próximo paso recomendado", ""]
    lines += [
        "Antes de construir CRUD completo, implementar una primera vertical chica:",
        "",
        "1. `CostingEngine` puro + tests.",
        "2. Schema Drizzle para `business_settings`, `suppliers`, `ingredients`, `recipes`, `recipe_lines`, `sub_recipes`, `sub_recipe_lines`.",
        "3. Importador read-only que produzca reporte, sin escribir DB todavía.",
        "4. Pantalla mínima de insumos/receta sólo para validar que el motor cierra.",
        "",
        "Esto evita el error clásico: UI preciosa con fundamentos financieros flojos. En costos gastronómicos, si el número está mal, TODO está mal.",
    ]
    lines += ["", "## Apéndice A — Fórmulas de muestra relevantes", ""]
    for sf in features["business_relevant_sample_formulas"][:40]:
        formula = str(sf.get("formula", "")).replace("\n", " ")
        lines.append(f"- `{sf.get('workbook')}` / `{sf.get('sheet')}` / `{sf.get('cell')}`: `{formula}`")
    return "\n".join(lines).rstrip() + "\n"


def infer_role(name):
    if name.startswith("Receta"):
        return "Ficha de receta individual"
    return {
        "Instrucciones": "Ayuda / onboarding",
        "Configuracion": "Parámetros globales del negocio",
        "Base de Datos": "Catálogo de insumos y proveedores",
        "Inventario": "Conteo físico / stock",
        "Pedidos": "Pedido sugerido",
        "Ingresos y Gastos": "Ledger mensual",
        "Punto de Equilibrio": "Reporte financiero",
        "Analisis de Precios": "Reporte de pricing",
        "Indice de Recetas": "Resumen derivado de recetas",
        "Sub-recetas": "Preparaciones reutilizables",
    }.get(name, "Pendiente")


def summarize_pdf_rules(pdf_audit):
    # Human-curated from extracted table of contents + keyword snippets; keep compact and implementation-oriented.
    return [
        ("Food cost", "No es ganancia; mide la proporción del precio que se consume en materia prima. Debe calcularse contra precio sin impuesto cuando corresponda."),
        ("Merma/factor de corrección", "Debe modelarse explícitamente porque afecta el costo real comprable/usable. No mezclarlo con consumo diario."),
        ("Margen de contribución", "No confundir margen bruto con plata disponible final; sirve para entender cuánto aporta cada venta a cubrir estructura."),
        ("Punto de equilibrio", "Debe responder cuántas unidades y ventas hacen falta por mes/día para no perder plata."),
        ("Costo teórico vs real", "El sistema debe permitir comparar cálculo estándar contra realidad operativa; las diferencias no son bug, son señal de gestión."),
        ("Estandarización", "Sin recetas fijas no hay costeo confiable. El editor de recetas debe empujar cantidades, unidades y rendimientos claros."),
        ("Sub-recetas", "Deben ser preparaciones reutilizables con rendimiento propio y costo por unidad de salida, no líneas mágicas."),
        ("Inventario y pedidos", "El pedido sugerido debe derivarse de stock, consumo y punto de reposición, manteniendo trazabilidad del motivo."),
        ("Errores comunes", "La app debe prevenir precios cargados a ojo, unidades inconsistentes, mermas omitidas y cálculos cacheados sin trazabilidad."),
    ]


def main():
    audits = [audit_excel(p) for p in EXCELS]
    pdf_audits = [extract_pdf(p) for p in PDFS]
    features = formula_features(audits)
    payload = {"excels": audits, "pdfs": pdf_audits, "formula_features": features}
    (OUT / "costeokit_source_audit.json").write_text(json.dumps(payload, indent=2, ensure_ascii=False, default=str), encoding="utf-8")
    md = build_markdown(audits, features, pdf_audits)
    (DATA / "COSTEOKIT_SOURCE_AUDIT.md").write_text(md, encoding="utf-8")
    print(DATA / "COSTEOKIT_SOURCE_AUDIT.md")
    print(OUT / "costeokit_source_audit.json")

if __name__ == "__main__":
    main()
